#!/usr/bin/env python3
"""
Convert Beechfield_Cap_Studio_Master_Data_v1.xlsx into the four JSON files
the prototype consumes. Run from the repo root:

    python3 cap-studio/scripts/convert_workbook.py /path/to/Beechfield_Cap_Studio_Master_Data_v1.xlsx

Outputs (overwritten):
    cap-studio/data/products.json
    cap-studio/data/filters.json
    cap-studio/data/scenarios.json
    cap-studio/data/field_definitions.json

No source values are renamed. Casing-only duplicates in Filter_Reference are
preserved so the UI can flag them as taxonomy QA issues.
"""
import json
import sys
from pathlib import Path

import openpyxl


def cell(v):
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if s == "" or s.lower() == "n/a":
            return None
        return s
    return v


def split_list(v, sep=";"):
    if v is None:
        return []
    if isinstance(v, (int, float)):
        return [v]
    return [p.strip() for p in str(v).split(sep) if p.strip()]


def main(xlsx_path: Path, out_dir: Path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # --- Product_Master ---
    pm = wb["Product_Master"]
    rows = list(pm.iter_rows(values_only=True))
    headers = list(rows[0])
    products = []
    for row in rows[1:]:
        rec = {}
        for h, v in zip(headers, row):
            rec[h] = cell(v)
        if rec.get("Colour List"):
            rec["Colour List Array"] = split_list(rec["Colour List"])
        else:
            rec["Colour List Array"] = []
        if rec.get("Size List"):
            rec["Size List Array"] = split_list(rec["Size List"])
        else:
            rec["Size List Array"] = []
        products.append(rec)

    # --- Filter_Reference ---
    fr = wb["Filter_Reference"]
    frows = list(fr.iter_rows(values_only=True))
    filters = {}
    for row in frows[1:]:
        group, value, source, notes = (cell(row[0]), cell(row[1]), cell(row[2]), cell(row[3]))
        if not group or value is None:
            continue
        filters.setdefault(group, []).append({"value": value, "source": source, "notes": notes})

    # Detect casing-only duplicates within each group
    for group, items in filters.items():
        seen = {}
        for it in items:
            key = str(it["value"]).strip().lower()
            seen.setdefault(key, []).append(it["value"])
        for key, vals in seen.items():
            uniq = list(dict.fromkeys(vals))
            if len(uniq) > 1:
                for it in items:
                    if str(it["value"]).strip().lower() == key:
                        it["casing_duplicate_of"] = uniq

    # --- Test_Scenarios ---
    ts = wb["Test_Scenarios"]
    trows = list(ts.iter_rows(values_only=True))
    th = list(trows[0])
    scenarios = []
    for row in trows[1:]:
        rec = {h: cell(v) for h, v in zip(th, row)}
        if not rec.get("Scenario ID"):
            continue
        rec["Market Filter(s) Array"] = split_list(rec.get("Market Filter(s)"))
        rec["Product Category Filter(s) Array"] = split_list(rec.get("Product Category Filter(s)"))
        rec["Product Type Filter(s) Array"] = split_list(rec.get("Product Type Filter(s)"))
        scenarios.append(rec)

    # --- Field_Definitions ---
    fd = wb["Field_Definitions"]
    fdrows = list(fd.iter_rows(values_only=True))
    fdh = list(fdrows[0])
    field_defs = []
    for row in fdrows[1:]:
        rec = {h: cell(v) for h, v in zip(fdh, row)}
        if rec.get("Field"):
            field_defs.append(rec)

    out_dir.mkdir(parents=True, exist_ok=True)
    (out_dir / "products.json").write_text(json.dumps(products, indent=2, ensure_ascii=False))
    (out_dir / "filters.json").write_text(json.dumps(filters, indent=2, ensure_ascii=False))
    (out_dir / "scenarios.json").write_text(json.dumps(scenarios, indent=2, ensure_ascii=False))
    (out_dir / "field_definitions.json").write_text(json.dumps(field_defs, indent=2, ensure_ascii=False))

    print(f"Wrote {len(products)} products, "
          f"{sum(len(v) for v in filters.values())} filter values across {len(filters)} groups, "
          f"{len(scenarios)} scenarios, "
          f"{len(field_defs)} field definitions "
          f"to {out_dir}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: convert_workbook.py <path-to-xlsx> [out-dir]", file=sys.stderr)
        sys.exit(1)
    xlsx = Path(sys.argv[1]).expanduser().resolve()
    out = Path(sys.argv[2]).expanduser().resolve() if len(sys.argv) > 2 else Path(__file__).resolve().parent.parent / "data"
    main(xlsx, out)
